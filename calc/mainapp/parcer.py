import os
from os import listdir
import openpyxl
from openpyxl import Workbook, load_workbook


mydir = '/Users/natalazivlova/Desktop/calc/calc/mainapp/'
myfile = 'excel.xlsm'
file = os.path.join(mydir, myfile)

wb = load_workbook(file, data_only=True)

print(f'Структура эксель файла {wb.sheetnames}')
print('------------------------------------------------------------------------------')

sheet = wb.worksheets[1]

control_actions_dict = {"name": [], "designation": [], "basic_equilibrium": [], "new_equilibrium": []}
for row in range(3, 13):
    name = sheet[row][0].value
    control_actions_dict['name'].append(name)
    designation = sheet[row][1].value
    control_actions_dict['designation'].append(designation)
    basic_equilibrium = sheet[row][2].value
    control_actions_dict['basic_equilibrium'].append(basic_equilibrium)
    new_equilibrium = sheet[row][3].value
    control_actions_dict['new_equilibrium'].append(new_equilibrium)
print(control_actions_dict)

variable = {"name": [], "designation": [], "basic_equilibrium": [], "basic_quantity": [], "relative_quality": [],
            "new_equilibrium": [], "new_quantity": [], "change_price": [], "change_quantity": [], "difference": []}
for row in range(18, 28):
    name = sheet[row][0].value
    variable['name'].append(name)
    designation = sheet[row][1].value
    variable['designation'].append(designation)
    basic_equilibrium = sheet[row][2].value
    variable['basic_equilibrium'].append(basic_equilibrium)
    basic_quantity = sheet[row][3].value
    variable['basic_quantity'].append(basic_quantity)
    relative_quality = sheet[row][4].value
    variable['relative_quality'].append(relative_quality)
    new_equilibrium = sheet[row][5].value
    variable['new_equilibrium'].append(new_equilibrium)
    new_quantity = sheet[row][6].value
    variable['new_quantity'].append(new_quantity)
    change_price = sheet[row][7].value
    variable['change_price'].append(change_price)
    change_quantity = sheet[row][8].value
    variable['change_quantity'].append(change_quantity)
    difference = sheet[row][9].value
    variable['difference'].append(difference)
print(variable)

equations = {"name": [], "formula": []}
for row in range(31, 51):
    name = sheet[row][0].value
    equations['name'].append(name)
    formula = sheet[row][1].value
    equations['formula'].append(formula)
print(equations)

# pretty = json.dumps(control_actions_dict)

behavioral_parameters = {"name": [], "meaning": []}
for row in range(2, 11):
    name = sheet[row][7].value
    behavioral_parameters['name'].append(name)
    meaning = sheet[row][8].value
    behavioral_parameters['meaning'].append(meaning)
print(behavioral_parameters)

behavioral_parameters_rho = {"name": [], "meaning": []}
for row in range(2, 5):
    name = sheet[row][9].value
    behavioral_parameters_rho['name'].append(name)
    meaning = sheet[row][10].value
    behavioral_parameters_rho['meaning'].append(meaning)
print(behavioral_parameters_rho)

behavioral_parameters_SS_DS = {"name": [], "meaning": []}
for row in range(6, 11):
    name = sheet[row][9].value
    behavioral_parameters_SS_DS['name'].append(name)
    meaning = sheet[row][10].value
    behavioral_parameters_SS_DS['meaning'].append(meaning)
print(behavioral_parameters_SS_DS)

external_values = {"name": [], "meaning": []}
for row in range(2, 9):
    name = sheet[row][12].value
    external_values['name'].append(name)
    meaning = sheet[row][13].value
    external_values['meaning'].append(meaning)
print(external_values)

taxes_welfare = {"name": [], "basic_equilibrium": [], "new_equilibrium": [], "change_million": [], "change_percent": []}
for row in range(18, 29):
    name = sheet[row][12].value
    taxes_welfare['name'].append(name)
    basic_equilibrium = sheet[row][13].value
    taxes_welfare['basic_equilibrium'].append(basic_equilibrium)
    new_equilibrium = sheet[row][14].value
    taxes_welfare['new_equilibrium'].append(new_equilibrium)
    change_million = sheet[row][15].value
    taxes_welfare['change_million'].append(change_million)
    change_percent = sheet[row][16].value
    taxes_welfare['change_percent'].append(change_percent)
print(taxes_welfare)

"""Лист Рентабельность (Росстат)"""

sheet = wb.worksheets[4]
profitability = {"classifier_forms": [], "classifier_types_OKVED2": [], "classifier_OKATO": [],
                 "classifier_org_forms": [], "unit": [], "period": [], "types_groupings": [], "y_2017": [],
                 "y_2018": [], "y_2019": [], "y_2020": []}
for row in range(2, 8):
    classifier_forms = sheet[row][0].value
    profitability['classifier_forms'].append(classifier_forms)
    classifier_types_OKVED2 = sheet[row][1].value
    profitability['classifier_types_OKVED2'].append(classifier_types_OKVED2)
    classifier_OKATO = sheet[row][2].value
    profitability['classifier_OKATO'].append(classifier_OKATO)
    classifier_org_forms = sheet[row][3].value
    profitability['classifier_org_forms'].append(classifier_org_forms)
    unit = sheet[row][4].value
    profitability['unit'].append(unit)
    period = sheet[row][5].value
    profitability['period'].append(period)
    types_groupings = sheet[row][6].value
    profitability['types_groupings'].append(types_groupings)
    y_2017 = sheet[row][7].value
    profitability['y_2017'].append(y_2017)
    y_2018 = sheet[row][8].value
    profitability['y_2018'].append(y_2018)
    y_2019 = sheet[row][9].value
    profitability['y_2019'].append(y_2019)
    y_2020 = sheet[row][10].value
    profitability['y_2020'].append(y_2020)
print(profitability)

"""Лист Эластичность спроса"""

sheet = wb.worksheets[3]
elasticity_demand = {"research": [], "price_elasticity": [], "link": []}
for row in range(2, 12):
    research = sheet[row][1].value
    elasticity_demand['research'].append(research)
    price_elasticity = sheet[row][2].value
    elasticity_demand['price_elasticity'].append(price_elasticity)
    link = sheet[row][3].value
    elasticity_demand['link'].append(link)
print(elasticity_demand)


"""Лист Структура затрат (ЦСР)"""

sheet = wb.worksheets[5]
cost_structure = {"cost_item": [], "rub_t": []}
for row in range(2, 15):
    cost_item = sheet[row][0].value
    cost_structure['cost_item'].append(cost_item)
    rub_t = sheet[row][1].value
    cost_structure['rub_t'].append(rub_t)
print(cost_structure)

"""Лист Эластичности замещения (GTAP)"""

sheet = wb.worksheets[6]
elasticity_substitution = {"number": [], "GTAP": [], "industry": [], "el_substitution_imp_dom": [],
                           "el_substitution_imp": []}
for row in range(2, 59):
    number = sheet[row][0].value
    elasticity_substitution['number'].append(number)
    GTAP = sheet[row][1].value
    elasticity_substitution['GTAP'].append(GTAP)
    industry = sheet[row][2].value
    elasticity_substitution['industry'].append(industry)
    el_substitution_imp_dom = sheet[row][3].value
    elasticity_substitution['el_substitution_imp_dom'].append(el_substitution_imp_dom)
    el_substitution_imp = sheet[row][4].value
    elasticity_substitution['el_substitution_imp'].append(el_substitution_imp)
print(elasticity_substitution)

"""Налоговая нагрузка (ФНС)"""

sheet = wb.worksheets[7]
tax_burden = {"type_ec_act_OKVED2": [], "tax": [], "fiscal_burden": []}
for row in range(6, 51):
    type_ec_act_OKVED2 = sheet[row][0].value
    tax_burden['type_ec_act_OKVED2'].append(type_ec_act_OKVED2)
    tax = sheet[row][1].value
    tax_burden['tax'].append(tax)
    fiscal_burden = sheet[row][2].value
    tax_burden['fiscal_burden'].append(fiscal_burden)
print(tax_burden)

""" Лист Рентабельность (ФНС) """
sheet = wb.worksheets[8]
profitability = {"type_act_OKVED2": [], "sold_goods": [], "assets": []}
for row in range(5, 66):
    type_act_OKVED2 = sheet[row][0].value
    profitability['type_act_OKVED2'].append(type_act_OKVED2)
    sold_goods = sheet[row][1].value
    profitability['sold_goods'].append(sold_goods)
    assets = sheet[row][2].value
    profitability['assets'].append(assets)
print(profitability)
